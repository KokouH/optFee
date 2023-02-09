from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl

from models import Chain


def get_if_exist(tag: str, di: dict):
    if di is None:
        return None
    if tag in di:
        return di[tag]
    return None


def create_tables_by_accounts(accounts, bookName):
    wb = openpyxl.Workbook()

    chains = [
        Chain.ChainName.ARBITRUM,
        Chain.ChainName.OPTIMISM,
        Chain.ChainName.ETHEREUM,
        Chain.ChainName.FANTOM,
        Chain.ChainName.POLYGON]

    arbitrum = wb.active
    arbitrum.title = 'Optimism'
    arbitrum.append(['Wallet', 'Transaction count', 'First trans UTC',
                     'Last trans UTC', 'AVG trans time', 'Fee(token)', 'Fee(USD)'])
    for cell in arbitrum['1']:
        cell.font = Font(color='ffffff', b=True, size=14)
        cell.fill = PatternFill('solid', fgColor='000000')
    chain_name = 'opt'
    line = 2
    side = Side(border_style="medium", color="000000")
    
    for ind, acc in enumerate(accounts):
        arbitrum.append([acc.wallet,
                         get_if_exist(chain_name, acc.trans_count),
                         get_if_exist(chain_name, acc.fist_trans),
                         get_if_exist(chain_name, acc.last_trans),
                         get_if_exist(chain_name, acc.avg_time_trans),
                         get_if_exist('count', get_if_exist(
                             chain_name, acc.fee_pay)),
                         get_if_exist('in_usd', get_if_exist(chain_name, acc.fee_pay))])

        for cell in arbitrum[str(line)]:
            if ind % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='d8d8d8')
            else:
                cell.fill = PatternFill('solid', fgColor='bfbfbf')
            cell.border = Border(right=side)
        line += 1

    arbitrum.merge_cells("K1:O1")
    arbitrum['K1'].font = Font(b=True, size=14)
    arbitrum['K1'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K1'] = 'ETH SPENT ON GAS'
    arbitrum.merge_cells('K2:O7')
    arbitrum['K2'].font = Font(b=True, size=20)
    arbitrum['K2'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K2'].fill = PatternFill('solid', fgColor='f2f2f2')
    arbitrum['K2'].border = Border(
        right=side, top=side, left=side, bottom=side)
    arbitrum['K2'] = '=SUM(F:F)'

    arbitrum.merge_cells("K9:O9")
    arbitrum['K9'].font = Font(b=True, size=14)
    arbitrum['K9'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K9'] = '$USD SPENT ON GAS (at current price)'
    arbitrum.merge_cells('K10:O15')
    arbitrum['K10'].font = Font(b=True, size=20)
    arbitrum['K10'].alignment = Alignment(
        horizontal='center', vertical='center')
    arbitrum['K10'].fill = PatternFill('solid', fgColor='f2f2f2')
    arbitrum['K10'].border = Border(
        right=side, top=side, left=side, bottom=side)
    arbitrum['K10'] = '=SUM(G:G)'

    arbitrum.column_dimensions['A'].width = 45
    arbitrum.column_dimensions['B'].width = 20
    arbitrum.column_dimensions['C'].width = 20
    arbitrum.column_dimensions['D'].width = 20
    arbitrum.column_dimensions['E'].width = 17
    arbitrum.column_dimensions['F'].width = 25
    arbitrum.column_dimensions['G'].width = 15


    arbitrumTokens = wb.create_sheet('Optimism tokens')
    for acc in accounts:
        arbitrumTokens.append([acc.wallet, 'Tokens', 'USD', '===NEW WALLET==='])
        tokens = get_if_exist(chain_name, acc.token_balance)
        if tokens is None:
            continue
        for token_name in tokens:
            tk = tokens[token_name]
            arbitrumTokens.append([f"{token_name} ( {get_if_exist('symbol', tk)} )", get_if_exist('count', tk), get_if_exist('in_usd', tk)])
    arbitrumTokens.column_dimensions['A'].width = 45
    arbitrumTokens.column_dimensions['B'].width = 20
    arbitrumTokens.column_dimensions['C'].width = 20

    # arbitrumNFTs = wb.create_sheet('Arbitrum NFTs')
    # for acc in accounts:
    #   arbitrumNFTs.append([acc.wallet, 'Name', 'Count', 'USD', '===NEW WALLET==='])
    #   chain_nfts = get_if_exist(chain_name, acc.nfts)
    #   if chain_nfts is None:
    #       continue
    #   for nft_name in chain_nfts:
    #       arbitrumNFTs.append([None, nft_name, chain_nfts[nft_name]['count'], get_if_exist('in_usd', chain_nfts[nft_name])])

   
    wb.save(str(bookName) + '.xlsx')
